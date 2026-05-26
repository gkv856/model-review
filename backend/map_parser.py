"""
Step 0 — Map File Parser
Input:  path to map .xlsx (sheet names + cell addresses mirror the model file)
Output: whitelist dict  { sheet_name: { cell_coord: symbol } }
Symbols: N=hardcoded, F=unique formula, S=sum, C=callup, X=external link
"""

import logging
import openpyxl

logger = logging.getLogger(__name__)

VALID_SYMBOLS = {"N", "F", "S", "C", "X"}


def parse_map(map_path: str) -> dict[str, dict[str, str]]:
    """
    Input:  path to map .xlsx file
    Output: { "P&L": { "F15": "F", "G20": "S" }, "DCF": { "D10": "X" }, ... }
    Raises: ValueError if no valid symbols found (map is empty or unrecognised)
    """
    wb = openpyxl.load_workbook(map_path, data_only=True)
    whitelist: dict[str, dict[str, str]] = {}
    total_symbols = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_cells: dict[str, str] = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value in VALID_SYMBOLS:
                    sheet_cells[cell.coordinate] = str(cell.value)

        symbol_count = len(sheet_cells)
        total_symbols += symbol_count
        whitelist[sheet_name] = sheet_cells

        log_data = {"sheet": sheet_name, "symbol_count": symbol_count}
        logger.info("[map_parser] Sheet parsed: %s", log_data)

    if total_symbols == 0:
        raise ValueError("Map file appears empty or invalid — no recognised symbols (N/F/S/C/X) found.")

    summary = {"total_symbols": total_symbols, "sheets": len(whitelist)}
    logger.info("[map_parser] Parse complete: %s", summary)

    return whitelist


def symbol_counts(whitelist: dict[str, dict[str, str]]) -> dict[str, int]:
    """
    Input:  whitelist from parse_map
    Output: count per symbol across all sheets  { "F": 48, "S": 15, ... }
    """
    counts: dict[str, int] = {s: 0 for s in VALID_SYMBOLS}
    for sheet_cells in whitelist.values():
        for symbol in sheet_cells.values():
            counts[symbol] = counts.get(symbol, 0) + 1
    return counts
