"""
Tests for map_parser.py
Builds in-memory xlsx workbooks so no fixture files are needed.
"""

import io
import pytest
import openpyxl

from map_parser import parse_map, symbol_counts, VALID_SYMBOLS


def _make_map_xlsx(sheet_data: dict[str, dict[str, str]]) -> str:
    """
    Helper — writes a map workbook to a temp file and returns the path.
    sheet_data: { sheet_name: { cell_coord: symbol_or_value } }
    """
    import tempfile, os

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for sheet_name, cells in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        for coord, value in cells.items():
            ws[coord] = value

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


class TestParseMap:
    def test_extracts_all_valid_symbols(self):
        path = _make_map_xlsx({
            "P&L": {"F15": "F", "F16": "N", "G15": "S", "H10": "C", "I10": "X"},
            "DCF": {"D10": "F", "D11": "S"},
        })
        result = parse_map(path)

        assert "P&L" in result
        assert "DCF" in result
        assert result["P&L"]["F15"] == "F"
        assert result["P&L"]["F16"] == "N"
        assert result["P&L"]["G15"] == "S"
        assert result["P&L"]["H10"] == "C"
        assert result["P&L"]["I10"] == "X"
        assert result["DCF"]["D10"] == "F"

    def test_ignores_empty_cells(self):
        path = _make_map_xlsx({"Sheet1": {"A1": "F", "A2": None, "A3": "S"}})
        result = parse_map(path)
        assert "A2" not in result["Sheet1"]
        assert len(result["Sheet1"]) == 2

    def test_ignores_unrecognised_values(self):
        # Cells with values like numbers or other strings are not symbols
        path = _make_map_xlsx({"Sheet1": {"A1": "F", "A2": 42, "A3": "Z", "A4": "S"}})
        result = parse_map(path)
        assert set(result["Sheet1"].values()) <= VALID_SYMBOLS
        assert len(result["Sheet1"]) == 2

    def test_raises_on_empty_map(self):
        path = _make_map_xlsx({"Sheet1": {"A1": None, "A2": None}})
        with pytest.raises(ValueError, match="empty or invalid"):
            parse_map(path)

    def test_multiple_sheets(self):
        path = _make_map_xlsx({
            "P&L":    {"F1": "F"},
            "Balance": {"G1": "S"},
            "DCF":    {"H1": "N"},
        })
        result = parse_map(path)
        assert len(result) == 3

    def test_empty_sheet_included_with_zero_count(self):
        path = _make_map_xlsx({
            "WithData": {"A1": "F"},
            "Empty":    {},
        })
        result = parse_map(path)
        assert "Empty" in result
        assert len(result["Empty"]) == 0


class TestSymbolCounts:
    def test_counts_all_symbols(self):
        whitelist = {
            "P&L": {"F15": "F", "F16": "N", "G15": "S"},
            "DCF": {"D10": "F", "D11": "C", "D12": "X"},
        }
        counts = symbol_counts(whitelist)
        assert counts["F"] == 2
        assert counts["N"] == 1
        assert counts["S"] == 1
        assert counts["C"] == 1
        assert counts["X"] == 1

    def test_zero_counts_for_missing_symbols(self):
        whitelist = {"Sheet1": {"A1": "F"}}
        counts = symbol_counts(whitelist)
        assert counts["N"] == 0
        assert counts["S"] == 0
        assert counts["C"] == 0
        assert counts["X"] == 0
