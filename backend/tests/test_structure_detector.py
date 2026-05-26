"""Tests for structure_detector.py"""

import tempfile
import openpyxl
import pytest

from core.structure_detector import detect_structure


def _make_wb(sheet_data: dict[str, dict[str, object]]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cells in sheet_data.items():
        ws = wb.create_sheet(name)
        for coord, val in cells.items():
            ws[coord] = val
    return wb


class TestTimelineDetection:
    def test_detects_year_integers(self):
        # Row 1: years 2020-2025 in columns B-G
        wb = _make_wb({"P&L": {
            "B1": 2020, "C1": 2021, "D1": 2022, "E1": 2023, "F1": 2024, "G1": 2025,
            "A2": "Revenue", "B2": 100,
        }})
        result = detect_structure(wb)
        assert result["P&L"]["timeline_row"] == 1

    def test_detects_fy_strings(self):
        wb = _make_wb({"Sheet1": {
            "B1": "FY2022", "C1": "FY2023", "D1": "FY2024", "E1": "FY2025",
        }})
        result = detect_structure(wb)
        assert result["Sheet1"]["timeline_row"] == 1

    def test_no_timeline_returns_none(self):
        wb = _make_wb({"Sheet1": {"A1": "Revenue", "B1": 100, "C1": 200}})
        result = detect_structure(wb)
        assert result["Sheet1"]["timeline_row"] is None


class TestLabelColDetection:
    def test_detects_string_heavy_column(self):
        # Column A rows 5-20 are mostly strings
        cells = {f"A{r}": f"Label {r}" for r in range(5, 22)}
        cells.update({f"B{r}": r * 10 for r in range(5, 22)})
        wb = _make_wb({"Sheet1": cells})
        result = detect_structure(wb)
        assert result["Sheet1"]["label_col"] == "A"

    def test_no_label_col_returns_none(self):
        # All numbers — no string column
        cells = {f"A{r}": r * 10 for r in range(5, 22)}
        wb = _make_wb({"Sheet1": cells})
        result = detect_structure(wb)
        assert result["Sheet1"]["label_col"] is None


class TestSectionHeaders:
    def test_detects_single_string_row(self):
        wb = _make_wb({"Sheet1": {
            "A1": "Revenue Build",  # lone string in row 1
            "A2": "Line item A", "B2": 100,
            "A3": "Line item B", "B3": 200,
        }})
        result = detect_structure(wb)
        headers = result["Sheet1"]["section_headers"]
        labels = [h["label"] for h in headers]
        assert "Revenue Build" in labels

    def test_section_headers_sorted_ascending(self):
        wb = _make_wb({"Sheet1": {
            "A10": "Section B",
            "A3":  "Section A",
        }})
        result = detect_structure(wb)
        rows = [h["row"] for h in result["Sheet1"]["section_headers"]]
        assert rows == sorted(rows)


class TestStructureMapKeys:
    def test_output_has_required_keys(self):
        wb = _make_wb({"Sheet1": {"A1": "x"}})
        result = detect_structure(wb)
        required = {"sheet", "timeline_row", "label_col", "unit_col", "data_start_col", "section_headers"}
        assert required.issubset(result["Sheet1"].keys())

    def test_all_sheets_present(self):
        wb = _make_wb({"P&L": {"A1": 1}, "DCF": {"A1": 2}, "BS": {"A1": 3}})
        result = detect_structure(wb)
        assert set(result.keys()) == {"P&L", "DCF", "BS"}
