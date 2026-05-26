"""
Tests for parser.py
Builds in-memory model + map workbooks — no fixture files needed.
"""

import tempfile
import openpyxl
import pytest

from core.map_parser import parse_map
from core.model_parser import parse_model


def _write_xlsx(sheet_data: dict[str, dict[str, object]]) -> str:
    """Helper: writes a workbook to a temp file, returns path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, cells in sheet_data.items():
        ws = wb.create_sheet(sheet_name)
        for coord, value in cells.items():
            ws[coord] = value
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    return tmp.name


class TestParseModel:
    def test_extracts_whitelisted_cells_only(self):
        # Model has A1..A5; map only marks A1 and A3
        model_path = _write_xlsx({"P&L": {"A1": 100, "A2": 200, "A3": 300, "A4": 400}})
        map_path   = _write_xlsx({"P&L": {"A1": "N", "A3": "F"}})

        whitelist = parse_map(map_path)
        cells     = parse_model(model_path, whitelist)

        coords = {c["cell"] for c in cells}
        assert coords == {"A1", "A3"}
        assert "A2" not in coords
        assert "A4" not in coords

    def test_cell_dict_has_required_keys(self):
        model_path = _write_xlsx({"Sheet1": {"B2": 42}})
        map_path   = _write_xlsx({"Sheet1": {"B2": "N"}})

        whitelist = parse_map(map_path)
        cells     = parse_model(model_path, whitelist)

        required_keys = {"sheet", "cell", "row", "col", "symbol", "formula", "value", "is_formula"}
        assert required_keys.issubset(cells[0].keys())

    def test_hardcoded_value_is_not_formula(self):
        model_path = _write_xlsx({"Sheet1": {"A1": 999}})
        map_path   = _write_xlsx({"Sheet1": {"A1": "N"}})

        whitelist = parse_map(map_path)
        cells     = parse_model(model_path, whitelist)

        cell = cells[0]
        assert cell["value"] == 999
        assert cell["is_formula"] is False
        assert cell["symbol"] == "N"

    def test_skips_sheet_missing_from_model(self):
        model_path = _write_xlsx({"Sheet1": {"A1": 1}})
        map_path   = _write_xlsx({"Sheet1": {"A1": "F"}, "MissingSheet": {"B1": "S"}})

        whitelist = parse_map(map_path)
        cells     = parse_model(model_path, whitelist)

        sheets = {c["sheet"] for c in cells}
        assert "MissingSheet" not in sheets
        assert "Sheet1" in sheets

    def test_row_and_col_populated(self):
        model_path = _write_xlsx({"Sheet1": {"C5": 10}})
        map_path   = _write_xlsx({"Sheet1": {"C5": "F"}})

        whitelist = parse_map(map_path)
        cells     = parse_model(model_path, whitelist)

        cell = cells[0]
        assert cell["row"] == 5
        assert cell["col"] == 3  # column C = 3

    def test_multi_sheet_extraction(self):
        model_path = _write_xlsx({
            "P&L": {"A1": 10, "B1": 20},
            "DCF": {"A1": 30},
        })
        map_path = _write_xlsx({
            "P&L": {"A1": "F", "B1": "S"},
            "DCF": {"A1": "C"},
        })

        whitelist = parse_map(map_path)
        cells     = parse_model(model_path, whitelist)

        assert len(cells) == 3
        sheets = {c["sheet"] for c in cells}
        assert sheets == {"P&L", "DCF"}
