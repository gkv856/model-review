"""
Shared pytest fixtures and factory helpers used across all backend tests.
Import helpers directly (e.g. `from tests.conftest import make_cell`)
or use fixtures via normal pytest injection.
"""

import tempfile
import pytest
import openpyxl
import networkx as nx


# ── workbook factory ──────────────────────────────────────────────────────────

def _build_wb(sheet_data: dict[str, dict[str, object]]) -> openpyxl.Workbook:
    """Build an in-memory workbook from a {sheet: {coord: value}} spec."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cells in sheet_data.items():
        ws = wb.create_sheet(name)
        for coord, val in cells.items():
            ws[coord] = val
    return wb


@pytest.fixture
def make_xlsx():
    """
    Fixture factory — returns a helper that writes a temp .xlsx and returns its path.
    Usage: path = make_xlsx({"Sheet1": {"A1": "F", "B1": 100}})
    """
    paths: list[str] = []

    def _make(sheet_data: dict[str, dict[str, object]]) -> str:
        wb  = _build_wb(sheet_data)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        paths.append(tmp.name)
        return tmp.name

    yield _make

    # Cleanup temp files after each test
    import os
    for p in paths:
        try:
            os.unlink(p)
        except OSError:
            pass


@pytest.fixture
def make_wb():
    """Fixture that returns the in-memory workbook builder (no file I/O)."""
    return _build_wb


# ── cell dict factory ─────────────────────────────────────────────────────────

def make_cell(
    *,
    sheet: str = "Sheet1",
    cell: str = "A1",
    symbol: str = "F",
    formula: str = "=B1+B2",
    value: object = 100,
    row: int = 1,
    col: int = 1,
    risk_score: float = 50.0,
    tier: int | str = 2,
    raw_dependencies: list | None = None,
) -> dict:
    """Return a minimal valid cell dict for unit tests."""
    return {
        "sheet":            sheet,
        "cell":             cell,
        "row":              row,
        "col":              col,
        "symbol":           symbol,
        "formula":          formula,
        "value":            value,
        "is_formula":       bool(formula and formula.startswith("=")),
        "risk_score":       risk_score,
        "tier":             tier,
        "raw_dependencies": raw_dependencies or [],
    }


# ── graph factory ─────────────────────────────────────────────────────────────

def make_graph(edges: list[tuple]) -> nx.DiGraph:
    """
    Build a DiGraph from a list of (source_node, target_node) edge tuples.
    Nodes are (sheet, cell) tuples: (("Sheet1", "A1"), ("Sheet1", "B1")).
    """
    G = nx.DiGraph()
    G.add_edges_from(edges)
    return G
