"""Tests for enricher.py"""

import openpyxl
import pytest

from tests.conftest import make_cell
from enricher import enrich_cell, enrich_cells, _find_section


def _make_wb(sheet_data: dict) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, cells in sheet_data.items():
        ws = wb.create_sheet(name)
        for coord, val in cells.items():
            ws[coord] = val
    return wb


def _base_structure(sheet: str = "Sheet1") -> dict:
    return {
        sheet: {
            "timeline_row": 1,
            "label_col":    "A",
            "unit_col":     "B",
            "data_start_col": "C",
            "section_headers": [
                {"row": 3, "label": "Revenue"},
                {"row": 8, "label": "Costs"},
            ],
        }
    }


class TestFindSection:
    def test_returns_correct_section(self):
        headers = [{"row": 3, "label": "Revenue"}, {"row": 8, "label": "Costs"}]
        assert _find_section(headers, row=10) == "Costs"
        assert _find_section(headers, row=5) == "Revenue"

    def test_falls_back_to_general(self):
        headers = [{"row": 5, "label": "Revenue"}]
        assert _find_section(headers, row=2) == "General"

    def test_empty_headers_returns_general(self):
        assert _find_section([], row=10) == "General"


class TestEnrichCell:
    def test_label_read_from_label_col(self):
        wb = _make_wb({"Sheet1": {"A5": "Total Revenue", "C5": 100}})
        cell = make_cell(cell="C5", row=5, col=3)
        enrich_cell(cell, _base_structure(), wb)
        assert cell["label"] == "Total Revenue"

    def test_label_fallback_when_empty(self):
        wb = _make_wb({"Sheet1": {"C5": 100}})
        cell = make_cell(cell="C5", row=5, col=3)
        enrich_cell(cell, _base_structure(), wb)
        assert cell["label"] == "Row 5"

    def test_units_read_from_unit_col(self):
        wb = _make_wb({"Sheet1": {"B5": "USD", "C5": 100}})
        cell = make_cell(cell="C5", row=5, col=3)
        enrich_cell(cell, _base_structure(), wb)
        assert cell["units"] == "USD"

    def test_period_read_from_timeline_row(self):
        # Timeline row is 1; cell is at col 3 → read C1
        wb = _make_wb({"Sheet1": {"C1": "FY2025", "C5": 100}})
        cell = make_cell(cell="C5", row=5, col=3)
        enrich_cell(cell, _base_structure(), wb)
        assert cell["period"] == "FY2025"

    def test_section_assigned_correctly(self):
        wb = _make_wb({"Sheet1": {"C10": 50}})
        cell = make_cell(cell="C10", row=10, col=3)
        enrich_cell(cell, _base_structure(), wb)
        assert cell["section"] == "Costs"   # row 10 ≥ header row 8

    def test_dependencies_populated(self):
        wb = _make_wb({
            "Sheet1": {
                "A2": "Input Label",
                "B2": 42,
                "C5": 100,
            }
        })
        cell = make_cell(cell="C5", row=5, col=3)
        cell["raw_dependencies"] = [("Sheet1", "B2")]
        enrich_cell(cell, _base_structure(), wb)
        assert len(cell["dependencies"]) == 1
        dep = cell["dependencies"][0]
        assert dep["cell"] == "B2"
        assert dep["value"] == 42
        assert dep["label"] == "Input Label"   # from A2 (label_col)

    def test_missing_sheet_graceful(self):
        wb = _make_wb({"Sheet1": {}})
        cell = make_cell(cell="C5", sheet="MissingSheet", row=5, col=3)
        enrich_cell(cell, {}, wb)   # no structure map entry
        assert cell["label"] == "Row 5"
        assert cell["section"] == "General"


class TestEnrichCells:
    def test_mutates_and_returns_same_list(self):
        wb = _make_wb({"Sheet1": {"C5": 100, "C6": 200}})
        cells = [
            make_cell(cell="C5", row=5, col=3),
            make_cell(cell="C6", row=6, col=3),
        ]
        result = enrich_cells(cells, _base_structure(), wb)
        assert result is cells
        for c in cells:
            assert "label" in c
            assert "section" in c
