"""
Step 2 — Structure Detector
Input:  loaded openpyxl workbook (full, not whitelist-filtered)
Output: structure map  { sheet_name: SheetStructure dict }
        Per sheet: timeline_row, label_col, unit_col, data_start_col, section_headers
Used only for enrichment context in Step 9 — not for formula review itself.
"""

import logging
import re
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils import col_letter

logger = logging.getLogger(__name__)

# Year range considered valid for timeline detection
TIMELINE_YEAR_MIN = 2015
TIMELINE_YEAR_MAX = 2045

TIMELINE_PATTERNS = re.compile(
    r"^(FY|CY)\d{4}$|^Q[1-4]\s?\d{4}$|^H[12]\s?\d{4}$|^\d{4}[EAF]$",
    re.IGNORECASE,
)

UNIT_STRINGS = {
    "usd", "mn", "bn", "%", "x", "inr", "days", "#", "bps",
    "gbp", "eur", "aud", "k", "m", "b", "thousands", "millions",
}


# ── helpers ────────────────────────────────────────────────────────────────


def _is_timeline_value(val: object) -> bool:
    """Returns True if the cell value looks like a year or period label."""
    if isinstance(val, int) and TIMELINE_YEAR_MIN <= val <= TIMELINE_YEAR_MAX:
        return True
    if isinstance(val, str) and TIMELINE_PATTERNS.match(val.strip()):
        return True
    return False


# col_letter imported from utils (shared utility)


# ── per-sheet detection ─────────────────────────────────────────────────────


def _detect_timeline_row(ws: Worksheet) -> int | None:
    """
    Scan first 10 rows — return the row number where >50% of non-empty
    cells are timeline values (years, FYxxxx, Qx xxxx, etc.).
    """
    for row_idx in range(1, 11):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row_vals if v is not None]
        if not non_empty:
            continue
        timeline_hits = sum(1 for v in non_empty if _is_timeline_value(v))
        if timeline_hits / len(non_empty) > 0.5:
            return row_idx
    return None


def _detect_label_col(ws: Worksheet) -> str | None:
    """
    Scan columns A–F — return the first column where >60% of cells in
    rows 5–200 are non-empty strings.
    """
    max_row = min(ws.max_row or 200, 200)
    scan_rows = range(5, max_row + 1)
    total = max(len(scan_rows), 1)

    for col_idx in range(1, 7):  # A=1 .. F=6
        string_count = 0
        for row_idx in scan_rows:
            val = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and val.strip():
                string_count += 1
        if string_count / total > 0.6:
            return col_letter(col_idx)
    return None


def _detect_unit_col(ws: Worksheet, label_col: str) -> str | None:
    """
    Check the column immediately right of label_col.
    Return it if >30% of cells in rows 5–200 contain known unit strings.
    """
    from openpyxl.utils import column_index_from_string
    label_idx = column_index_from_string(label_col)
    unit_idx  = label_idx + 1

    max_row    = min(ws.max_row or 200, 200)
    scan_rows  = range(5, max_row + 1)
    total      = max(len(scan_rows), 1)
    unit_count = 0

    for row_idx in scan_rows:
        val = ws.cell(row=row_idx, column=unit_idx).value
        if isinstance(val, str) and val.strip().lower() in UNIT_STRINGS:
            unit_count += 1

    if unit_count / total > 0.3:
        return col_letter(unit_idx)
    return None


def _detect_section_headers(ws: Worksheet) -> list[dict]:
    """
    Return rows where a single non-empty string exists and >80% of other
    cells in the row are empty — or a merged cell spans >3 columns.
    Output: [{ "row": int, "label": str }] sorted ascending.
    """
    headers: list[dict] = []
    max_col = max(ws.max_column or 1, 1)

    # Build merged cell set for quick lookup
    merged_ranges = list(ws.merged_cells.ranges)

    for row_idx in range(1, (ws.max_row or 1) + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        non_empty = [(i, v) for i, v in enumerate(row_vals) if v is not None]

        # Single string cell with 80%+ empty neighbours
        if len(non_empty) == 1:
            _, val = non_empty[0]
            if isinstance(val, str) and val.strip():
                headers.append({"row": row_idx, "label": val.strip()})
                continue

        # Merged cell spanning >3 columns in this row
        for mr in merged_ranges:
            if mr.min_row == row_idx and (mr.max_col - mr.min_col) >= 3:
                val = ws.cell(row=row_idx, column=mr.min_col).value
                if isinstance(val, str) and val.strip():
                    headers.append({"row": row_idx, "label": val.strip()})
                    break

    return sorted(headers, key=lambda h: h["row"])


# ── public API ───────────────────────────────────────────────────────────────


def detect_structure(wb: Workbook) -> dict[str, dict]:
    """
    Input:  fully loaded openpyxl Workbook (data_only=True recommended)
    Output: { sheet_name: { timeline_row, label_col, unit_col,
                             data_start_col, section_headers } }
    Falls back gracefully if any detection step fails.
    """
    structure_map: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        try:
            timeline_row = _detect_timeline_row(ws)
        except Exception:
            timeline_row = None

        try:
            label_col = _detect_label_col(ws)
        except Exception:
            label_col = None

        try:
            unit_col = _detect_unit_col(ws, label_col) if label_col else None
        except Exception:
            unit_col = None

        try:
            section_headers = _detect_section_headers(ws)
        except Exception:
            section_headers = []

        # data_start_col is the column after label + optional unit column
        if label_col:
            from openpyxl.utils import column_index_from_string
            offset = 2 if unit_col else 1
            data_start_col = col_letter(column_index_from_string(label_col) + offset)
        else:
            data_start_col = None

        structure_map[sheet_name] = {
            "sheet":           sheet_name,
            "timeline_row":    timeline_row,
            "label_col":       label_col,
            "unit_col":        unit_col,
            "data_start_col":  data_start_col,
            "section_headers": section_headers,
        }

        log_data = {
            "sheet":        sheet_name,
            "timeline_row": timeline_row,
            "label_col":    label_col,
            "unit_col":     unit_col,
            "sections":     len(section_headers),
        }
        logger.info("[structure_detector] %s", log_data)

    return structure_map
