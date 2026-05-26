"""
Step 0 — Map File Parser
Input:  path to map .xlsx
Output: whitelist { sheet_name: { cell_coord: symbol } }
"""

import openpyxl

from utils.logger import get_logger

logger = get_logger(__name__)

VALID_SYMBOLS = {"N", "F", "S", "C", "X"}


def parse_map(map_path: str) -> dict[str, dict[str, str]]:
    """
    Input:  path to map .xlsx
    Output: { "P&L": { "F15": "F", "G20": "S" }, ... }
    Raises: ValueError if no valid symbols found
    """
    wb = openpyxl.load_workbook(map_path, data_only=True)
    whitelist: dict[str, dict[str, str]] = {}
    total = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_cells: dict[str, str] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in VALID_SYMBOLS:
                    sheet_cells[cell.coordinate] = str(cell.value)
        total += len(sheet_cells)
        whitelist[sheet_name] = sheet_cells
        logger.info("[map_parser] Sheet parsed: %s", {"sheet": sheet_name, "symbols": len(sheet_cells)})

    if total == 0:
        raise ValueError("Map file appears empty or invalid — no recognised symbols (N/F/S/C/X) found.")

    logger.info("[map_parser] Complete: %s", {"total_symbols": total, "sheets": len(whitelist)})
    return whitelist


def symbol_counts(whitelist: dict[str, dict[str, str]]) -> dict[str, int]:
    """Input: whitelist. Output: count per symbol across all sheets."""
    counts: dict[str, int] = {s: 0 for s in VALID_SYMBOLS}
    for sheet_cells in whitelist.values():
        for sym in sheet_cells.values():
            counts[sym] = counts.get(sym, 0) + 1
    return counts
