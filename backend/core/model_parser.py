"""
Step 1 — Model File Parser
Input:  model .xlsx path + whitelist from map_parser
Output: list of cell dicts for whitelisted cells only
"""

import openpyxl

from utils.logger import get_logger

logger = get_logger(__name__)


def parse_model(model_path: str, whitelist: dict[str, dict[str, str]]) -> list[dict]:
    """
    Input:  model_path, whitelist from parse_map()
    Output: list of cell dicts — sheet, cell, row, col, symbol, formula, value, is_formula
    Loads workbook twice: formulas (data_only=False) and values (data_only=True).
    """
    wb_formulas = openpyxl.load_workbook(model_path, data_only=False)
    wb_values   = openpyxl.load_workbook(model_path, data_only=True)

    cells: list[dict] = []
    skipped: list[str] = []

    for sheet_name, cell_map in whitelist.items():
        if sheet_name not in wb_formulas.sheetnames:
            skipped.append(sheet_name)
            logger.warning("[model_parser] Sheet in map not in model — skipping: %s", sheet_name)
            continue

        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        for coord, symbol in cell_map.items():
            cf = ws_f[coord]
            cv = ws_v[coord]
            formula_str = str(cf.value) if cf.value is not None else None
            cells.append({
                "sheet":      sheet_name,
                "cell":       coord,
                "row":        cf.row,
                "col":        cf.column,
                "symbol":     symbol,
                "formula":    formula_str,
                "value":      cv.value,
                "is_formula": bool(formula_str and formula_str.startswith("=")),
            })

    logger.info("[model_parser] Parsed: %s", {
        "cells": len(cells),
        "sheets_ok": len(whitelist) - len(skipped),
        "skipped": skipped,
    })
    return cells
