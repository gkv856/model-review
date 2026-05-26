"""
Step 2 — Structure Detector
Input:  loaded openpyxl workbook
Output: structure map { sheet_name: { timeline_row, label_col, unit_col, data_start_col, section_headers } }

Parallelised per sheet using ThreadPoolExecutor at 65% CPU capacity.
"""

import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import os

from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from utils.helpers import col_letter
from utils.logger import get_logger

logger = get_logger(__name__)

_MAX_WORKERS = max(1, int((os.cpu_count() or 1) * 0.65))

TIMELINE_YEAR_MIN = 2015
TIMELINE_YEAR_MAX = 2045

TIMELINE_PATTERNS = re.compile(
    r"^(FY|CY)\d{4}$|^Q[1-4]\s?\d{4}$|^H[12]\s?\d{4}$|^\d{4}[EAF]$",
    re.IGNORECASE,
)

UNIT_STRINGS = {
    "usd", "mn", "bn", "%", "x", "inr", "days", "#", "bps",
    "gbp", "eur", "aud", "k", "m", "b", "thousands", "millions",
}


def _is_timeline_value(val: object) -> bool:
    if isinstance(val, int) and TIMELINE_YEAR_MIN <= val <= TIMELINE_YEAR_MAX:
        return True
    if isinstance(val, str) and TIMELINE_PATTERNS.match(val.strip()):
        return True
    return False


def _detect_timeline_row(ws: Worksheet) -> int | None:
    for row_idx in range(1, 11):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in row_vals if v is not None]
        if not non_empty:
            continue
        if sum(1 for v in non_empty if _is_timeline_value(v)) / len(non_empty) > 0.5:
            return row_idx
    return None


def _detect_label_col(ws: Worksheet) -> str | None:
    max_row = min(ws.max_row or 200, 200)
    scan_rows = range(5, max_row + 1)
    total = max(len(scan_rows), 1)
    for col_idx in range(1, 7):
        hits = sum(
            1 for r in scan_rows
            if isinstance(ws.cell(row=r, column=col_idx).value, str)
            and ws.cell(row=r, column=col_idx).value.strip()
        )
        if hits / total > 0.6:
            return col_letter(col_idx)
    return None


def _detect_unit_col(ws: Worksheet, label_col: str) -> str | None:
    unit_idx = column_index_from_string(label_col) + 1
    max_row = min(ws.max_row or 200, 200)
    scan_rows = range(5, max_row + 1)
    total = max(len(scan_rows), 1)
    hits = sum(
        1 for r in scan_rows
        if isinstance(ws.cell(row=r, column=unit_idx).value, str)
        and ws.cell(row=r, column=unit_idx).value.strip().lower() in UNIT_STRINGS
    )
    return col_letter(unit_idx) if hits / total > 0.3 else None


def _detect_section_headers(ws: Worksheet) -> list[dict]:
    headers: list[dict] = []
    max_col = max(ws.max_column or 1, 1)
    merged_ranges = list(ws.merged_cells.ranges)
    for row_idx in range(1, (ws.max_row or 1) + 1):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, max_col + 1)]
        non_empty = [(i, v) for i, v in enumerate(row_vals) if v is not None]
        if len(non_empty) == 1:
            _, val = non_empty[0]
            if isinstance(val, str) and val.strip():
                headers.append({"row": row_idx, "label": val.strip()})
                continue
        for mr in merged_ranges:
            if mr.min_row == row_idx and (mr.max_col - mr.min_col) >= 3:
                val = ws.cell(row=row_idx, column=mr.min_col).value
                if isinstance(val, str) and val.strip():
                    headers.append({"row": row_idx, "label": val.strip()})
                    break
    return sorted(headers, key=lambda h: h["row"])


def _detect_sheet(sheet_name: str, ws: Worksheet) -> tuple[str, dict]:
    """Detect structure for a single sheet — designed to run in a thread."""
    try:
        timeline_row = _detect_timeline_row(ws)
    except Exception:
        timeline_row = None
    try:
        label_col = _detect_label_col(ws)
    except Exception:
        label_col = None
    try:
        unit_col = _detect_unit_col(ws, label_col) if label_col else None
    except Exception:
        unit_col = None
    try:
        section_headers = _detect_section_headers(ws)
    except Exception:
        section_headers = []

    if label_col:
        offset = 2 if unit_col else 1
        data_start_col = col_letter(column_index_from_string(label_col) + offset)
    else:
        data_start_col = None

    logger.info("[structure_detector] %s", {
        "sheet": sheet_name, "timeline_row": timeline_row,
        "label_col": label_col, "sections": len(section_headers),
    })
    return sheet_name, {
        "sheet":           sheet_name,
        "timeline_row":    timeline_row,
        "label_col":       label_col,
        "unit_col":        unit_col,
        "data_start_col":  data_start_col,
        "section_headers": section_headers,
    }


def detect_structure(wb: Workbook) -> dict[str, dict]:
    """
    Input:  openpyxl Workbook
    Output: structure map per sheet

    Sheets are processed in parallel (up to 65% CPU workers).
    """
    sheets = wb.sheetnames
    structure_map: dict[str, dict] = {}

    workers = min(_MAX_WORKERS, len(sheets))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {
            pool.submit(_detect_sheet, name, wb[name]): name
            for name in sheets
        }
        for future in as_completed(futures):
            name, result = future.result()
            structure_map[name] = result

    # Preserve workbook sheet order
    return {name: structure_map[name] for name in sheets if name in structure_map}
