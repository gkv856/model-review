"""
Step 1 — Model File Parser
Input:  model .xlsx path + whitelist from map_parser
Output: list of cell dicts for whitelisted cells only
        Each dict: sheet, cell, row, col, symbol, formula, value, is_formula
"""

import logging
import openpyxl

logger = logging.getLogger(__name__)


def parse_model(model_path: str, whitelist: dict[str, dict[str, str]]) -> list[dict]:
    """
    Input:  model_path — path to model .xlsx
            whitelist  — { sheet: { coord: symbol } } from parse_map()
    Output: list of cell dicts, one per whitelisted cell
    Loads workbook twice: once for formulas, once for computed values.
    Sheets in whitelist but missing from model are logged and skipped.
    """
    wb_formulas = openpyxl.load_workbook(model_path, data_only=False)
    wb_values   = openpyxl.load_workbook(model_path, data_only=True)

    cells: list[dict] = []
    skipped_sheets: list[str] = []

    for sheet_name, cell_map in whitelist.items():
        if sheet_name not in wb_formulas.sheetnames:
            skipped_sheets.append(sheet_name)
            logger.warning("[parser] Sheet in map not found in model — skipping: %s", sheet_name)
            continue

        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name]

        for coord, symbol in cell_map.items():
            cf = ws_f[coord]
            cv = ws_v[coord]

            raw_formula = cf.value
            formula_str = str(raw_formula) if raw_formula is not None else None
            is_formula  = bool(formula_str and formula_str.startswith("="))

            cells.append({
                "sheet":      sheet_name,
                "cell":       coord,
                "row":        cf.row,
                "col":        cf.column,
                "symbol":     symbol,
                "formula":    formula_str,
                "value":      cv.value,
                "is_formula": is_formula,
            })

    summary = {
        "total_cells_extracted": len(cells),
        "sheets_processed": len(whitelist) - len(skipped_sheets),
        "sheets_skipped": skipped_sheets,
    }
    logger.info("[parser] Model parsed: %s", summary)

    return cells
