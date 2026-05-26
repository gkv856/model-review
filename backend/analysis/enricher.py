"""
Step 9 — Context Enricher
Input:  deduplicated representative cell dicts + structure map + data_only workbook
Output: same cells with label, units, period, section, dependencies added (mutates in place)

Run only on Tier 1 + Tier 2 representatives.
"""

from openpyxl.utils.cell import coordinate_to_tuple

from utils.logger import get_logger

logger = get_logger(__name__)


def _get_val(ws, col_letter: str | None, row: int):
    """Read a single cell value given a column letter and row."""
    if not col_letter:
        return None
    try:
        return ws[f"{col_letter}{row}"].value
    except Exception:
        return None


def _get_val_by_rowcol(ws, row: int, col: int):
    """Read a single cell value given row and column integers."""
    try:
        return ws.cell(row=row, column=col).value
    except Exception:
        return None


def _find_section(section_headers: list[dict], row: int) -> str:
    """
    Walk section_headers in reverse order.
    Return the label of the first header whose row <= cell's row.
    Falls back to 'General'.
    """
    for header in reversed(section_headers):
        if header["row"] <= row:
            return header["label"]
    return "General"


def enrich_cell(cell: dict, structure_map: dict, wb_values) -> dict:
    """
    Input:  single cell dict, structure map from structure_detector, data_only workbook
    Output: same dict with label, units, period, section, dependencies added

    Mutates cell in place and returns it.
    """
    sheet = cell["sheet"]
    row   = cell["row"]
    col   = cell["col"]

    sm = structure_map.get(sheet, {})
    ws = wb_values[sheet] if sheet in wb_values.sheetnames else None

    label_col = sm.get("label_col")
    label = _get_val(ws, label_col, row) if ws else None
    cell["label"] = str(label) if label is not None else f"Row {row}"

    unit_col = sm.get("unit_col")
    cell["units"] = _get_val(ws, unit_col, row) if (ws and unit_col) else None

    timeline_row = sm.get("timeline_row")
    if ws and timeline_row:
        cell["period"] = _get_val_by_rowcol(ws, timeline_row, col)
    else:
        cell["period"] = None

    section_headers = sm.get("section_headers", [])
    cell["section"] = _find_section(section_headers, row)

    dep_details = []
    for dep_sheet, dep_coord in cell.get("raw_dependencies", []):
        dep_row, _ = coordinate_to_tuple(dep_coord)
        dep_sm = structure_map.get(dep_sheet, {})
        dep_label_col = dep_sm.get("label_col")

        dep_label = None
        dep_value = None

        if dep_sheet in wb_values.sheetnames:
            dep_ws = wb_values[dep_sheet]
            dep_label = _get_val(dep_ws, dep_label_col, dep_row)
            dep_value = dep_ws[dep_coord].value if dep_coord else None

        dep_details.append({
            "sheet": dep_sheet,
            "cell":  dep_coord,
            "label": str(dep_label) if dep_label is not None else dep_coord,
            "value": dep_value,
        })

    cell["dependencies"] = dep_details
    return cell


def enrich_cells(cells: list[dict], structure_map: dict, wb_values) -> list[dict]:
    """
    Input:  list of cell dicts (Tier 1 + Tier 2 representatives)
    Output: same list with enrichment added to each cell (mutates in place)
    """
    for cell in cells:
        enrich_cell(cell, structure_map, wb_values)

    logger.info("[enricher] Enriched %d cells", len(cells))
    return cells
